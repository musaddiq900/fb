import os
import re
import time
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urljoin
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, Response
from flask_socketio import SocketIO, emit
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, Workbook
import pandas as pd
import json
import queue

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Global state
scraping_status = {
    'is_running': False,
    'processed': 0,
    'total': 0,
    'successful': 0,
    'failed': 0,
    'current_url': '',
    'progress': 0,
    'speed': 0
}

data_queue = queue.Queue()
scraped_data = []
EXCEL_FILENAME = 'facebook_pages_data.xlsx'

# Configure patterns for data detection
EMAIL_PATTERN = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

PHONE_PATTERNS = [
    r'(?:\+92|0092|92)?[-\s]?(3\d{2})[-\s]?(\d{7})',
    r'(0?3\d{2})[-\s]?(\d{7})',
    r'(0?3\d{2})\s?(\d{3}\s?\d{4})',
    r'(\(0?3\d{2}\))\s?(\d{7})',
    r'(?:\+971|00971|971)?[-\s]?(\d)[-\s]?(\d{3})[-\s]?(\d{4})',
]

COMPILED_PATTERNS = {
    'email': re.compile(EMAIL_PATTERN, re.IGNORECASE),
    'phone': [re.compile(pattern) for pattern in PHONE_PATTERNS],
    'website': re.compile(r'(https?://[^\s]+|www\.[^\s]+)'),
    'likes': re.compile(r'(\d+[,.]?\d*)\s*(likes|people\s+like\s+this)', re.IGNORECASE),
    'followers': re.compile(r'(\d+[,.]?\d*)\s*(followers|people\s+follow\s+this)', re.IGNORECASE)
}

# Thread-local storage for WebDriver
thread_local = threading.local()

def get_driver():
    """Get or create a WebDriver instance"""
    if not hasattr(thread_local, 'driver'):
        chrome_options = Options()
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        ]
        
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument(f'user-agent={random.choice(user_agents)}')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        try:
            driver = webdriver.Chrome(options=chrome_options)
            driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
            })
            thread_local.driver = driver
        except Exception as e:
            print(f"Error creating driver: {e}")
            return None
    
    return thread_local.driver

def close_driver():
    """Close WebDriver instance"""
    if hasattr(thread_local, 'driver'):
        thread_local.driver.quit()
        del thread_local.driver

def extract_using_patterns(text, pattern_type):
    """Extract data using pre-compiled regex patterns"""
    if not text:
        return ""
    
    if pattern_type == 'email':
        match = COMPILED_PATTERNS['email'].search(text)
        return match.group(0) if match else ""
    
    elif pattern_type == 'phone':
        for pattern in COMPILED_PATTERNS['phone']:
            match = pattern.search(text)
            if match:
                phone = ''.join([g for g in match.groups() if g])
                return normalize_phone(phone)
        return ""
    
    elif pattern_type == 'website':
        match = COMPILED_PATTERNS['website'].search(text)
        return match.group(0) if match else ""
    
    return ""

def normalize_phone(phone):
    """Normalize phone number"""
    if not phone:
        return ""
    cleaned = re.sub(r'[^\d+]', '', phone)
    return cleaned if len(cleaned) >= 8 else ""

def detect_country(phone):
    """Detect country from phone number"""
    if phone.startswith('92') or phone.startswith('+92'):
        return "Pakistan"
    elif phone.startswith('971') or phone.startswith('+971'):
        return "UAE"
    elif phone.startswith('1') or phone.startswith('+1'):
        return "US/Canada"
    return "Unknown"

def save_to_excel(data, filename=EXCEL_FILENAME):
    """Save data to Excel file"""
    lock = threading.Lock()
    with lock:
        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
                if ws.max_row == 1 and ws['A1'].value != 'Name':
                    headers = ['Name', 'Email', 'Phone', 'Country', 'Page Link', 'Website', 
                              'Location', 'Address', 'Likes', 'Followers', 'Scrape Time']
                    ws.append(headers)
            else:
                wb = Workbook()
                ws = wb.active
                headers = ['Name', 'Email', 'Phone', 'Country', 'Page Link', 'Website', 
                          'Location', 'Address', 'Likes', 'Followers', 'Scrape Time']
                ws.append(headers)
            
            ws.append([
                data.get('name', '')[:100],
                data.get('email', ''),
                data.get('phone', ''),
                data.get('country', ''),
                data.get('page_link', ''),
                data.get('website', ''),
                data.get('location', ''),
                data.get('address', ''),
                data.get('likes', ''),
                data.get('followers', ''),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ])
            
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False

def scrape_facebook_page(url):
    """Scrape a single Facebook page"""
    driver = get_driver()
    if not driver:
        return None
    
    data = {
        'name': '', 'email': '', 'phone': '', 'country': '',
        'page_link': url, 'website': '', 'location': '',
        'address': '', 'likes': '', 'followers': ''
    }
    
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        page_text = driver.find_element(By.TAG_NAME, 'body').text
        
        # Extract name
        try:
            name_elem = driver.find_element(By.XPATH, "//h1 | //title")
            data['name'] = name_elem.text.strip()[:100]
        except:
            data['name'] = url.split('/')[-1]
        
        # Extract contact info
        data['email'] = extract_using_patterns(page_text, 'email')
        phone_num = extract_using_patterns(page_text, 'phone')
        data['phone'] = phone_num
        data['country'] = detect_country(phone_num) if phone_num else ""
        data['website'] = extract_using_patterns(page_text, 'website')
        
        # Extract social stats
        likes_match = COMPILED_PATTERNS['likes'].search(page_text)
        followers_match = COMPILED_PATTERNS['followers'].search(page_text)
        
        data['likes'] = likes_match.group(1).replace(',', '') if likes_match else ""
        data['followers'] = followers_match.group(1).replace(',', '') if followers_match else ""
        
        # Try to get location/address
        try:
            location_elem = driver.find_elements(By.XPATH, 
                "//div[contains(text(), 'Location') or contains(text(), 'Address')]/following-sibling::div")
            if location_elem:
                data['address'] = location_elem[0].text[:200]
        except:
            pass
        
        # Save to Excel
        save_to_excel(data)
        
        return data
        
    except Exception as e:
        print(f"Error scraping {url}: {e}")
        return None
    finally:
        driver.get("about:blank")



def extract_links_from_html(html_content, base_url='https://www.facebook.com'):
    """Extract Facebook links from HTML content"""
    soup = BeautifulSoup(html_content, 'html.parser')
    links = set()
    
    # Look for Facebook links in various formats
    facebook_patterns = [
        r'https?://(www\.)?facebook\.com/[A-Za-z0-9_.-]+/?',
        r'https?://(www\.)?facebook\.com/pages/[^/]+/\d+',
        r'https?://(www\.)?facebook\.com/profile\.php\?id=\d+',
        r'https?://(www\.)?facebook\.com/groups/[A-Za-z0-9_.-]+'
    ]
    
    for a_tag in soup.find_all('a', href=True):
        href = a_tag['href']
        for pattern in facebook_patterns:
            if re.match(pattern, href):
                if base_url and not href.startswith('http'):
                    href = urljoin(base_url, href)
                links.add(href)
                break
    
    return list(links)

def update_scraping_status(status_update):
    """Update scraping status and emit via SocketIO"""
    global scraping_status
    scraping_status.update(status_update)
    socketio.emit('status_update', scraping_status)

def scraping_worker(links, html_content=None):
    """Main scraping worker function"""
    global scraping_status
    
    all_links = []
    
    # If HTML content is provided, extract links from it
    if html_content:
        extracted_links = extract_links_from_html(html_content)
        all_links.extend(extracted_links)
    
    # Add any direct links
    all_links.extend(links)
    
    if not all_links:
        update_scraping_status({'is_running': False})
        return
    
    total = len(all_links)
    update_scraping_status({
        'is_running': True,
        'total': total,
        'processed': 0,
        'successful': 0,
        'failed': 0,
        'progress': 0
    })
    
    # Process links with threading
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = []
        for link in all_links:
            futures.append(executor.submit(process_single_url, link))
        
        for future in as_completed(futures):
            try:
                result = future.result()
                if result:
                    data_queue.put(result)
                    socketio.emit('new_data', result)
                    
                    # Update status
                    scraping_status['processed'] += 1
                    scraping_status['successful'] += 1
                else:
                    scraping_status['failed'] += 1
                    scraping_status['processed'] += 1
                
                # Calculate progress
                progress = (scraping_status['processed'] / total) * 100
                scraping_status['progress'] = round(progress, 1)
                
                # Emit status update
                socketio.emit('status_update', scraping_status)
                
            except Exception as e:
                print(f"Error in future: {e}")
                scraping_status['failed'] += 1
                scraping_status['processed'] += 1
    
    # Final cleanup
    update_scraping_status({'is_running': False})
    close_driver()

def process_single_url(url):
    """Process a single URL and return data"""
    update_scraping_status({'current_url': url})
    
    start_time = time.time()
    data = scrape_facebook_page(url)
    elapsed = time.time() - start_time
    
    if data:
        data['scrape_time'] = round(elapsed, 2)
        return data
    return None

# Flask Routes
@app.route('/')
def index():
    """Render main page"""
    return render_template('index.html')

@app.route('/api/scrape', methods=['POST'])
def start_scraping():
    """Start scraping process"""
    global scraping_status
    
    if scraping_status['is_running']:
        return jsonify({'error': 'Scraping already in progress'}), 400
    
    data = request.json
    links = data.get('links', [])
    html_content = data.get('html_content', '')
    
    # Start scraping in background thread
    thread = threading.Thread(target=scraping_worker, args=(links, html_content))
    thread.daemon = True
    thread.start()
    
    return jsonify({'message': 'Scraping started', 'status': 'processing'})

@app.route('/api/stop', methods=['POST'])
def stop_scraping():
    """Stop scraping process"""
    # Note: This is a simple implementation. For proper stopping,
    # you would need to implement thread termination logic
    update_scraping_status({'is_running': False})
    return jsonify({'message': 'Stop command sent'})

@app.route('/api/status')
def get_status():
    """Get current scraping status"""
    return jsonify(scraping_status)

@app.route('/api/data')
def get_data():
    """Get scraped data"""
    # Read data from Excel file
    try:
        if os.path.exists(EXCEL_FILENAME):
            df = pd.read_excel(EXCEL_FILENAME)
            data = df.to_dict('records')
            return jsonify(data)
        return jsonify([])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export')
def export_data():
    """Export data as Excel file"""
    if not os.path.exists(EXCEL_FILENAME):
        return jsonify({'error': 'No data available'}), 404
    
    return send_file(
        EXCEL_FILENAME,
        as_attachment=True,
        download_name=f'facebook_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/api/clear', methods=['POST'])
def clear_data():
    """Clear all scraped data"""
    try:
        if os.path.exists(EXCEL_FILENAME):
            os.remove(EXCEL_FILENAME)
        
        # Create fresh Excel file with headers
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Email', 'Phone', 'Country', 'Page Link', 'Website', 
                  'Location', 'Address', 'Likes', 'Followers', 'Scrape Time']
        ws.append(headers)
        wb.save(EXCEL_FILENAME)
        
        return jsonify({'message': 'Data cleared successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload_html', methods=['POST'])
def upload_html():
    """Upload and process HTML file"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and file.filename.endswith('.html'):
        html_content = file.read().decode('utf-8')
        links = extract_links_from_html(html_content)
        
        return jsonify({
            'message': 'File processed successfully',
            'links_found': len(links),
            'links': links[:50]  # Return first 50 links
        })
    
    return jsonify({'error': 'Invalid file format'}), 400

# SocketIO Events
@socketio.on('connect')
def handle_connect():
    """Handle client connection"""
    print('Client connected')
    emit('status_update', scraping_status)

@socketio.on('disconnect')
def handle_disconnect():
    """Handle client disconnection"""
    print('Client disconnected')

if __name__ == '__main__':
    # Create Excel file with headers if it doesn't exist
    if not os.path.exists(EXCEL_FILENAME):
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Email', 'Phone', 'Country', 'Page Link', 'Website', 
                  'Location', 'Address', 'Likes', 'Followers', 'Scrape Time']
        ws.append(headers)
        wb.save(EXCEL_FILENAME)
    
    print("Starting Facebook Page Scraper UI...")
    print("Open your browser and go to: http://localhost:5000")
    socketio.run(app, debug=True, port=5000)